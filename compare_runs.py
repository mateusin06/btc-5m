#!/usr/bin/env python3
"""
Ferramenta de backtesting - testa múltiplas configs e gera Excel.
"""

import argparse
import time
from dataclasses import dataclass
from typing import Optional

from backtest import fetch_candles_range, get_candles_for_window
from strategy import analyze

# Modelo de preço do token (igual ao bot)
def delta_to_token_price(delta_pct: float) -> float:
    abs_d = abs(delta_pct)
    if abs_d < 0.005:
        return 0.50
    if abs_d < 0.02:
        return 0.50 + (abs_d - 0.005) / 0.015 * 0.05
    if abs_d < 0.05:
        return 0.55 + (abs_d - 0.02) / 0.03 * 0.10
    if abs_d < 0.10:
        return 0.65 + (abs_d - 0.05) / 0.05 * 0.15
    if abs_d < 0.15:
        return 0.80 + (abs_d - 0.10) / 0.05 * 0.12
    return min(0.92 + (abs_d - 0.15) * 0.5, 0.97)


@dataclass
class BacktestConfig:
    min_confidence: float
    bet_pct: float
    use_flat_bet: bool  # True = flat 25%, False = mode-based


def run_backtest(
    candles: list[dict],
    start_ts: int,
    end_ts: int,
    config: BacktestConfig,
    initial_bankroll: float = 100.0,
) -> tuple[list[dict], float, int, int]:
    """
    Executa backtest. Retorna (trades, bankroll_final, wins, losses).
    """
    trades = []
    bankroll = initial_bankroll
    wins = 0
    losses = 0

    # Janelas 5min no intervalo
    ts = start_ts
    while ts + 300 <= end_ts:
        window_candles = get_candles_for_window(candles, ts, minutes_before=25)
        if len(window_candles) < 5:
            ts += 300
            continue

        # Preço de abertura = candle que começa em ts
        window_open = None
        for c in window_candles:
            if c["t"] // 1000 == ts:
                window_open = c["o"]
                break
        if window_open is None:
            ts += 300
            continue

        # Preço no "T-10s" = último candle antes do fechamento
        # Simulamos: usamos o candle que termina em ts+300
        close_candle = None
        for c in window_candles:
            if c["t"] // 1000 >= ts + 4 * 60:
                close_candle = c
                break
        if not close_candle:
            close_candle = window_candles[-1]
        price_at_snipe = close_candle["c"]

        # Resultado real: close >= open?
        close_price = None
        for c in window_candles:
            if c["t"] // 1000 == ts + 5 * 60 or (c["t"] // 1000 >= ts + 4 * 60 and c["t"] // 1000 < ts + 6 * 60):
                close_price = c["c"]
                break
        if close_price is None:
            close_price = window_candles[-1]["c"] if window_candles else price_at_snipe

        up_wins = close_price >= window_open

        # Análise
        result = analyze(window_open, price_at_snipe, window_candles, None)
        if result.confidence < config.min_confidence:
            ts += 300
            continue

        if config.use_flat_bet:
            bet = initial_bankroll * 0.25
        elif config.bet_pct >= 1.0:  # aggressive
            bet = bankroll - initial_bankroll if bankroll > initial_bankroll else bankroll
        else:
            bet = bankroll * config.bet_pct
        bet = min(bet, bankroll)
        if bet < 5:
            ts += 300
            continue

        token_price = delta_to_token_price(result.window_delta_pct)
        shares = bet / token_price
        won = (result.direction == "up" and up_wins) or (result.direction == "down" and not up_wins)
        pnl = (shares * 1.0 - bet) if won else -bet
        bankroll += pnl

        trades.append({
            "ts": ts,
            "direction": result.direction,
            "confidence": result.confidence,
            "delta_pct": result.window_delta_pct,
            "token_price": token_price,
            "bet": bet,
            "won": won,
            "pnl": pnl,
            "bankroll": bankroll,
        })
        if won:
            wins += 1
        else:
            losses += 1

        ts += 300

    return trades, bankroll, wins, losses


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--hours", type=int, default=72)
    parser.add_argument("--output", default="results.xlsx")
    args = parser.parse_args()

    end_ts = int(time.time()) // 300 * 300
    start_ts = end_ts - args.hours * 3600

    print(f"Baixando candles {args.hours}h...")
    candles = fetch_candles_range(start_ts=start_ts * 1000, end_ts=end_ts * 1000)
    print(f"  {len(candles)} candles")

    configs = []
    for min_conf in [0.0, 0.15, 0.20, 0.25, 0.30, 0.35, 0.40, 0.45, 0.50]:
        for mode in ["flat", "safe", "aggressive"]:
            configs.append(BacktestConfig(
                min_confidence=min_conf,
                bet_pct=0.25 if mode == "safe" else 1.0,
                use_flat_bet=(mode == "flat"),
            ))

    results = []
    for i, cfg in enumerate(configs):
        trades, final_br, wins, losses = run_backtest(candles, start_ts, end_ts, cfg)
        total = wins + losses
        wr = wins / total * 100 if total else 0
        results.append({
            "min_conf": cfg.min_confidence,
            "mode": "flat" if cfg.use_flat_bet else ("safe" if cfg.bet_pct == 0.25 else "aggressive"),
            "trades": total,
            "wins": wins,
            "losses": losses,
            "win_rate": wr,
            "final_bankroll": final_br,
            "roi": (final_br - 100) / 100 * 100,
        })
        if (i + 1) % 10 == 0:
            print(f"  {i + 1}/{len(configs)} configs...")

    # Melhor config
    best = max(results, key=lambda r: r["final_bankroll"])
    best_trades, _, _, _ = run_backtest(
        candles, start_ts, end_ts,
        BacktestConfig(min_confidence=best["min_conf"], bet_pct=0.25 if best["mode"] == "safe" else 1.0, use_flat_bet=(best["mode"] == "flat")),
    )

    try:
        import openpyxl
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.title = "Summary"
        ws.append(["Min Conf", "Mode", "Trades", "Wins", "Losses", "Win Rate %", "Final Bankroll", "ROI %"])
        for r in results:
            ws.append([r["min_conf"], r["mode"], r["trades"], r["wins"], r["losses"], round(r["win_rate"], 1), round(r["final_bankroll"], 2), round(r["roi"], 1)])

        ws2 = wb.create_sheet("Best Config Trades")
        ws2.append(["Timestamp", "Direction", "Confidence", "Delta %", "Token Price", "Bet", "Won", "PnL", "Bankroll"])
        for t in best_trades:
            ws2.append([t["ts"], t["direction"], round(t["confidence"], 3), round(t["delta_pct"], 4), round(t["token_price"], 2), round(t["bet"], 2), t["won"], round(t["pnl"], 2), round(t["bankroll"], 2)])

        ws3 = wb.create_sheet("Bankroll Curves")
        ws3.append(["Config", "Trades", "Final Bankroll"])
        for r in results[:10]:
            ws3.append([f"conf={r['min_conf']} mode={r['mode']}", r["trades"], round(r["final_bankroll"], 2)])

        wb.save(args.output)
        print(f"\nSalvo em {args.output}")
        print(f"Melhor: min_conf={best['min_conf']} mode={best['mode']} -> ${best['final_bankroll']:.2f} ({best['roi']:.1f}% ROI)")
    except ImportError:
        print("Instale openpyxl: pip install openpyxl")
        for r in results[:5]:
            print(r)


if __name__ == "__main__":
    main()
